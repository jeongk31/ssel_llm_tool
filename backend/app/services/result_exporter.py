"""Prepare communication episodes and build researcher-facing result datasets."""

from __future__ import annotations

import json
import math
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_MAX_CELL_CHARACTERS = 32_767


@dataclass(frozen=True)
class PreparedCodingDataset:
    """Episode table plus the episode assigned to every original source row."""

    episodes: pd.DataFrame
    source_episode_indices: list[int]


def _is_missing_scalar(value: Any) -> bool:
    """Return whether a dataset cell is missing without accepting array results."""

    try:
        missing = pd.isna(value)
        return not hasattr(missing, "__len__") and bool(missing)
    except (TypeError, ValueError):
        return False


def _typed_value_key(value: Any) -> tuple[str, Any]:
    """Use exact, type-aware equality while treating all missing cells alike."""

    if _is_missing_scalar(value):
        return ("missing", None)
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    if isinstance(value, pd.Timestamp):
        return ("timestamp", value.isoformat())
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return (type(value).__name__, value)
    return (type(value).__name__, str(value))


def _display_dataset_value(value: Any) -> str:
    if _is_missing_scalar(value) or (isinstance(value, str) and value.strip() == ""):
        return "(blank)"
    return str(value)


def find_context_conflicts(
    df: pd.DataFrame,
    *,
    identifier_columns: list[str] | None,
    context_columns: list[str] | None,
) -> list[dict[str, Any]]:
    """Find context fields that are not exactly constant within an episode.

    Missing values form one explicit value class, so a blank/nonblank mixture is
    a conflict. With no identifier columns, every row is its own episode and no
    within-episode conflict is possible.
    """

    id_columns = list(identifier_columns or [])
    selected_context = list(dict.fromkeys(context_columns or []))
    if not id_columns or not selected_context:
        return []

    missing_columns = [
        column for column in [*id_columns, *selected_context] if column not in df.columns
    ]
    if missing_columns:
        raise ValueError(
            "Mapped columns were not found in the uploaded dataset: "
            + ", ".join(missing_columns)
            + "."
        )

    grouped = list(df.groupby(id_columns, sort=False, dropna=False))
    conflicts: list[dict[str, Any]] = []
    for context_column in selected_context:
        conflict_count = 0
        example_episode: dict[str, str] | None = None
        example_values: list[str] | None = None
        for group_key, group in grouped:
            distinct: dict[tuple[str, Any], Any] = {}
            for value in group[context_column].tolist():
                distinct.setdefault(_typed_value_key(value), value)
            if len(distinct) <= 1:
                continue
            conflict_count += 1
            if example_episode is None:
                keys = group_key if isinstance(group_key, tuple) else (group_key,)
                example_episode = {
                    column: _display_dataset_value(value)
                    for column, value in zip(id_columns, keys)
                }
                example_values = [_display_dataset_value(value) for value in distinct.values()]

        if conflict_count:
            conflicts.append(
                {
                    "column": context_column,
                    "conflicting_episode_count": conflict_count,
                    "example_episode": example_episode or {},
                    "example_values": example_values or [],
                }
            )
    return conflicts


def context_conflict_message(conflicts: list[dict[str, Any]]) -> str:
    """Create a concise server-side error for bypassed UI validation."""

    details: list[str] = []
    for conflict in conflicts:
        episode = ", ".join(
            f"{column}={value}"
            for column, value in conflict.get("example_episode", {}).items()
        ) or "example episode"
        values = ", ".join(conflict.get("example_values", []))
        details.append(
            f"{conflict['column']}: {conflict['conflicting_episode_count']} "
            f"episode(s) conflict; {episode} contains {values}"
        )
    return (
        "Selected Context fields must contain exactly one value within every episode. "
        + "; ".join(details)
        + ". Correct and re-upload the dataset or remove the inconsistent Context field."
    )


def detect_sender_names(
    df: pd.DataFrame,
    identity_column: str | None,
) -> tuple[list[str], list[int]]:
    """Return distinct nonblank sender names and one-based blank row numbers."""

    if not identity_column or identity_column not in df.columns:
        return [], []
    names: list[str] = []
    seen: set[str] = set()
    blank_rows: list[int] = []
    for position, value in enumerate(df[identity_column].tolist(), start=1):
        if _is_missing_scalar(value) or str(value).strip() == "":
            blank_rows.append(position)
            continue
        name = str(value).strip()
        if name not in seen:
            seen.add(name)
            names.append(name)
    return names, blank_rows


def validate_sender_configuration(
    df: pd.DataFrame,
    *,
    identity_column: str | None,
    participants: list[str] | None,
    codebook: list[dict[str, Any]],
) -> list[str]:
    """Enforce that sender-level output keys match the mapped dataset exactly."""

    if not any(entry.get("level") == "sender" for entry in codebook):
        return list(participants or [])
    if not identity_column or identity_column not in df.columns:
        raise ValueError("Per-sender categories require a mapped Sender column.")
    detected, blank_rows = detect_sender_names(df, identity_column)
    if blank_rows:
        preview = ", ".join(str(row) for row in blank_rows[:5])
        suffix = "…" if len(blank_rows) > 5 else ""
        raise ValueError(
            f"Sender column '{identity_column}' contains blank values in source row(s) "
            f"{preview}{suffix}. Fill every sender value before using per-sender categories."
        )
    if not detected:
        raise ValueError(
            f"Sender column '{identity_column}' does not contain any nonblank sender names."
        )
    supplied = [str(value).strip() for value in (participants or []) if str(value).strip()]
    if supplied != detected:
        raise ValueError(
            "The sender list does not match the values detected from the mapped Sender "
            f"column. Expected, in first-appearance order: {', '.join(detected)}."
        )
    return detected


def prepare_coding_dataset(
    df: pd.DataFrame,
    *,
    message_column: str,
    identifier_columns: list[str] | None,
    identity_column: str | None,
    order_column: str | None,
    order_direction: str,
    context_columns: list[str] | None = None,
) -> PreparedCodingDataset:
    """Create the episode-level coding table without losing source-row lineage.

    Episode order and message construction intentionally match the application's
    existing preprocessing behavior.  ``source_episode_indices`` is positional:
    item *i* is the episode index for original source row *i*.
    """

    original = df.reset_index(drop=True).copy()
    id_columns = [c for c in (identifier_columns or []) if c in original.columns]
    conflicts = find_context_conflicts(
        original,
        identifier_columns=id_columns,
        context_columns=context_columns,
    )
    if conflicts:
        raise ValueError(context_conflict_message(conflicts))
    if not id_columns:
        return PreparedCodingDataset(
            episodes=original,
            source_episode_indices=list(range(len(original))),
        )

    marker = "__chat_source_row_position__"
    while marker in original.columns:
        marker = "_" + marker

    work = original.copy()
    work[marker] = list(range(len(work)))
    if order_column and order_column in work.columns:
        work = work.sort_values(
            by=order_column,
            ascending=(order_direction != "desc"),
            kind="stable",
        )

    use_identity = bool(identity_column) and identity_column in work.columns
    source_episode_indices = [-1] * len(original)
    units: list[dict[str, Any]] = []

    for episode_index, (_, group) in enumerate(
        work.groupby(id_columns, sort=False, dropna=False)
    ):
        parts: list[str] = []
        for _, row in group.iterrows():
            message = "" if pd.isna(row[message_column]) else str(row[message_column])
            who = row[identity_column] if use_identity else None
            # Do not turn a genuinely empty message into non-empty "[sender] "
            # text.  This keeps empty-message handling meaningful after grouping.
            if (
                message.strip()
                and use_identity
                and pd.notna(who)
                and str(who).strip()
            ):
                parts.append(f"[{who}] {message}")
            else:
                parts.append(message)

        unit = {column: group.iloc[0][column] for column in original.columns}
        unit[message_column] = "\n".join(parts)
        # These mapped fields describe individual source messages.  Preserve
        # their full, message-aligned sequences in the episode table instead of
        # presenting only the first sender or order value.
        for mapped_column in (identity_column, order_column):
            if (
                mapped_column
                and mapped_column in original.columns
                and mapped_column != message_column
                and mapped_column not in id_columns
            ):
                unit[mapped_column] = "\n".join(
                    "" if pd.isna(value) else str(value)
                    for value in group[mapped_column].tolist()
                )
        units.append(unit)

        for source_position in group[marker].tolist():
            source_episode_indices[int(source_position)] = episode_index

    if any(index < 0 for index in source_episode_indices):
        raise ValueError("Could not map every source row to a communication episode.")

    episode_df = pd.DataFrame(units, columns=list(original.columns)).reset_index(drop=True)
    return PreparedCodingDataset(
        episodes=episode_df,
        source_episode_indices=source_episode_indices,
    )


def expanded_codebook_labels(
    codebook: list[dict[str, Any]], participants: list[str] | None
) -> list[str]:
    """Return unique output keys in the same order used by the coding runner."""

    participant_names = participants or []
    labels: list[str] = []
    for entry in codebook:
        label = str(entry.get("label") or "").strip()
        if not label:
            continue
        if entry.get("level") == "sender" and participant_names:
            labels.extend(f"{label}_{participant}" for participant in participant_names)
        else:
            labels.append(label)

    seen: set[str] = set()
    duplicates: list[str] = []
    for label in labels:
        if label in seen and label not in duplicates:
            duplicates.append(label)
        seen.add(label)
    if duplicates:
        raise ValueError(
            "Codebook output labels must be unique. Rename the conflicting "
            f"variable or participant labels: {', '.join(duplicates)}."
        )
    return labels


def _unique_result_names(original_columns: list[Any], labels: list[str]) -> dict[str, str]:
    """Keep codebook labels unless one would overwrite an original column."""

    used = {str(column) for column in original_columns}
    names: dict[str, str] = {}
    for label in labels:
        candidate = label
        if candidate in used:
            candidate = f"{label}_coded"
        suffix = 2
        while candidate in used:
            candidate = f"{label}_coded_{suffix}"
            suffix += 1
        used.add(candidate)
        names[label] = candidate
    return names


def build_result_frames(
    *,
    source_df: pd.DataFrame,
    prepared: PreparedCodingDataset,
    coded_rows: list[dict[str, Any]],
    codebook: list[dict[str, Any]],
    participants: list[str] | None,
    message_column: str,
    identifier_columns: list[str] | None,
    context_columns: list[str] | None,
    identity_column: str | None = None,
    order_column: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str]]:
    """Build source-row and compact episode-level result tables.

    The most recent entry for an episode index wins.  This makes the function
    defensive to duplicate input while matching selective-rerun semantics in
    the browser, where a rerun replaces the prior aggregate for that episode.
    """

    labels = expanded_codebook_labels(codebook, participants)
    result_names = _unique_result_names(list(source_df.columns), labels)
    episode_results: dict[int, dict[str, Any]] = {}

    for item in coded_rows:
        index = item.get("index")
        if isinstance(index, bool) or not isinstance(index, int):
            raise ValueError("Every coded result must have an integer episode index.")
        if index < 0 or index >= len(prepared.episodes):
            raise ValueError(f"Coded result episode index {index} is out of range.")
        coded = item.get("coded")
        if not isinstance(coded, dict):
            raise ValueError(f"Coded result for episode {index} is not an object.")
        episode_results[index] = coded

    source_result = source_df.reset_index(drop=True).copy()
    for label in labels:
        destination = result_names[label]
        source_result[destination] = [
            episode_results.get(episode_index, {}).get(label)
            for episode_index in prepared.source_episode_indices
        ]

    selected_episode_columns: list[str] = []
    for column in [
        *(identifier_columns or []),
        message_column,
        identity_column,
        order_column,
        *(context_columns or []),
    ]:
        if not column:
            continue
        if column in prepared.episodes.columns and column not in selected_episode_columns:
            selected_episode_columns.append(column)

    episode_result = prepared.episodes.loc[:, selected_episode_columns].copy()
    for label in labels:
        destination = result_names[label]
        episode_result[destination] = [
            episode_results.get(index, {}).get(label)
            for index in range(len(prepared.episodes))
        ]

    return source_result, episode_result, result_names


def _excel_value(value: Any) -> Any:
    """Convert pandas/provider values to safe scalar Excel cell values."""

    if value is None:
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (ValueError, TypeError):
            pass
    try:
        missing = pd.isna(value)
        if not hasattr(missing, "__len__") and bool(missing):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.isoformat()
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    return str(value)


def _first_xml_illegal_character(value: str) -> str | None:
    """Return the first character that XML 1.0 cannot represent, if any."""

    for character in value:
        codepoint = ord(character)
        if (
            codepoint in (0x09, 0x0A, 0x0D)
            or 0x20 <= codepoint <= 0xD7FF
            or 0xE000 <= codepoint <= 0xFFFD
            or 0x10000 <= codepoint <= 0x10FFFF
        ):
            continue
        return character
    return None


def _validate_excel_string(value: str, *, sheet_name: str, coordinate: str) -> None:
    """Reject text that XLSX cannot preserve exactly instead of mutating it."""

    if len(value) > EXCEL_MAX_CELL_CHARACTERS:
        raise ValueError(
            f"Worksheet '{sheet_name}' cell {coordinate} contains {len(value):,} "
            f"characters; Excel supports at most {EXCEL_MAX_CELL_CHARACTERS:,} "
            "characters in one cell. Shorten this value before exporting."
        )

    illegal_character = _first_xml_illegal_character(value)
    if illegal_character is not None:
        raise ValueError(
            f"Worksheet '{sheet_name}' cell {coordinate} contains an "
            "XML-incompatible character "
            f"(U+{ord(illegal_character):04X}). Remove that character before "
            "exporting to XLSX."
        )


def _write_excel_cell(
    worksheet: Any,
    *,
    row: int,
    column: int,
    value: Any,
    sheet_name: str,
) -> None:
    """Write a validated value, keeping every string literal rather than a formula."""

    if isinstance(value, str):
        coordinate = f"{get_column_letter(column)}{row}"
        _validate_excel_string(
            value,
            sheet_name=sheet_name,
            coordinate=coordinate,
        )

    cell = worksheet.cell(row=row, column=column, value=value)
    # openpyxl interprets strings beginning with '=' as formulas.  Exported
    # headers and dataset values are research data, so every string is written
    # explicitly as a literal cell value.
    if isinstance(value, str):
        cell.data_type = "s"


def dataframes_to_xlsx(sheets: list[tuple[str, pd.DataFrame]]) -> bytes:
    """Write flat, analysis-friendly worksheets with modest usability styling."""

    if not sheets:
        raise ValueError("At least one worksheet is required.")
    for sheet_name, frame in sheets:
        if len(frame) + 1 > 1_048_576:
            raise ValueError(
                f"Worksheet '{sheet_name}' exceeds Excel's 1,048,576-row limit."
            )
        if len(frame.columns) > 16_384:
            raise ValueError(
                f"Worksheet '{sheet_name}' exceeds Excel's 16,384-column limit."
            )
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()
    for requested_name, df in sheets:
        sheet_name = requested_name[:31] or "Sheet"
        candidate = sheet_name
        suffix = 2
        while candidate in used_names:
            ending = f"_{suffix}"
            candidate = sheet_name[: 31 - len(ending)] + ending
            suffix += 1
        used_names.add(candidate)

        worksheet = workbook.create_sheet(candidate)
        worksheet.freeze_panes = "A2"

        headers = [str(column) for column in df.columns]
        for column_number, header in enumerate(headers, start=1):
            _write_excel_cell(
                worksheet,
                row=1,
                column=column_number,
                value=header,
                sheet_name=candidate,
            )
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for values in df.itertuples(index=False, name=None):
            row_number = worksheet.max_row + 1
            for column_number, value in enumerate(values, start=1):
                _write_excel_cell(
                    worksheet,
                    row=row_number,
                    column=column_number,
                    value=_excel_value(value),
                    sheet_name=candidate,
                )

        if headers:
            worksheet.auto_filter.ref = worksheet.dimensions

        # Size columns from a bounded sample so large research datasets remain fast.
        sample_end = min(worksheet.max_row, 501)
        for column_number, header in enumerate(headers, start=1):
            max_length = len(header)
            for row_number in range(2, sample_end + 1):
                value = worksheet.cell(row=row_number, column=column_number).value
                if value is not None:
                    parts = str(value).splitlines() or [""]
                    max_length = max(
                        max_length,
                        max(len(part) for part in parts),
                    )
            worksheet.column_dimensions[get_column_letter(column_number)].width = min(
                max(max_length + 2, 10),
                70 if "message" in header.lower() else 40,
            )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def dataframe_to_xlsx(df: pd.DataFrame, *, sheet_name: str) -> bytes:
    """Write one flat worksheet."""

    return dataframes_to_xlsx([(sheet_name, df)])


def dataframe_to_csv(df: pd.DataFrame) -> bytes:
    """Write a flat UTF-8 CSV while preserving the DataFrame's row and column order."""

    return df.to_csv(index=False, lineterminator="\n").encode("utf-8")
