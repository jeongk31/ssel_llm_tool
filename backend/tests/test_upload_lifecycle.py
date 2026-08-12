import asyncio
import io
import json
import os
import tempfile
import time
import unittest
import uuid
import zipfile
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from fastapi import FastAPI
from fastapi.testclient import TestClient
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded

from app.ratelimit import limiter
from app.routes import agreement, coding
from app.services import category_runner


class UploadLifecycleTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.tempdir_patch = patch.object(
            coding.tempfile,
            "gettempdir",
            return_value=self.temp_dir.name,
        )
        self.tempdir_patch.start()
        coding._uploaded_files.clear()

        app = FastAPI()
        app.state.limiter = limiter
        app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
        app.include_router(coding.router, prefix="/api")
        app.include_router(agreement.router, prefix="/api")
        self.client = TestClient(app)

    def tearDown(self):
        self.client.close()
        coding._uploaded_files.clear()
        self.tempdir_patch.stop()
        self.temp_dir.cleanup()

    def _store_csv(self) -> tuple[str, dict]:
        return coding._store_uploaded_file(
            b"message,round\nhello,1\n",
            "sample.csv",
            "csv",
        )

    def test_resolver_recovers_upload_after_process_cache_is_cleared(self):
        file_id, original = self._store_csv()
        coding._uploaded_files.clear()

        recovered = coding.resolve_uploaded_file(file_id)

        self.assertEqual(recovered["path"], original["path"])
        self.assertEqual(recovered["filename"], "sample.csv")
        self.assertEqual(
            Path(recovered["path"]).read_bytes(),
            b"message,round\nhello,1\n",
        )
        self.assertEqual(coding._uploaded_files[file_id], recovered)

    def test_resolver_expires_at_exact_24_hour_boundary(self):
        file_id, info = self._store_csv()
        coding._uploaded_files.clear()

        with self.assertRaises(coding.UploadResolutionError) as caught:
            coding.resolve_uploaded_file(
                file_id,
                now=info["created"] + coding._TEMP_TTL_SECONDS,
            )

        self.assertEqual(caught.exception.status_code, 410)
        self.assertEqual(caught.exception.code, "UPLOAD_EXPIRED")
        self.assertFalse(os.path.exists(coding._upload_dir_for_id(file_id)))

    def test_missing_metadata_does_not_fall_back_to_new_format_cache(self):
        file_id, _ = self._store_csv()
        metadata = Path(coding._upload_dir_for_id(file_id)) / coding._UPLOAD_METADATA_FILENAME
        metadata.unlink()

        with self.assertRaises(coding.UploadResolutionError) as caught:
            coding.resolve_uploaded_file(file_id)

        self.assertEqual(caught.exception.status_code, 410)
        self.assertEqual(caught.exception.code, "UPLOAD_GONE")
        self.assertFalse(os.path.exists(coding._upload_dir_for_id(file_id)))

    def test_missing_physical_dataset_is_reported_gone(self):
        file_id, info = self._store_csv()
        Path(info["path"]).unlink()
        coding._uploaded_files.clear()

        with self.assertRaises(coding.UploadResolutionError) as caught:
            coding.resolve_uploaded_file(file_id)

        self.assertEqual(caught.exception.code, "UPLOAD_GONE")
        self.assertFalse(os.path.exists(coding._upload_dir_for_id(file_id)))

    def test_metadata_path_traversal_is_rejected_without_deleting_target(self):
        file_id, _ = self._store_csv()
        upload_dir = Path(coding._upload_dir_for_id(file_id))
        metadata_path = upload_dir / coding._UPLOAD_METADATA_FILENAME
        outside = Path(self.temp_dir.name) / "outside.csv"
        outside.write_text("message\nkeep me\n", encoding="utf-8")
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        metadata["stored_filename"] = "../outside.csv"
        metadata_path.write_text(json.dumps(metadata), encoding="utf-8")
        coding._uploaded_files.clear()

        with self.assertRaises(coding.UploadResolutionError) as caught:
            coding.resolve_uploaded_file(file_id)

        self.assertEqual(caught.exception.code, "UPLOAD_GONE")
        self.assertTrue(outside.is_file())
        self.assertFalse(upload_dir.exists())

    def test_nonfinite_and_far_future_creation_times_are_rejected(self):
        invalid_times = (
            float("nan"),
            float("inf"),
            time.time() + coding._UPLOAD_MAX_FUTURE_SKEW_SECONDS + 60,
        )
        for invalid_time in invalid_times:
            with self.subTest(created=invalid_time):
                file_id, _ = self._store_csv()
                upload_dir = Path(coding._upload_dir_for_id(file_id))
                metadata_path = upload_dir / coding._UPLOAD_METADATA_FILENAME
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
                metadata["created"] = invalid_time
                metadata_path.write_text(json.dumps(metadata), encoding="utf-8")
                coding._uploaded_files.clear()

                with self.assertRaises(coding.UploadResolutionError) as caught:
                    coding.resolve_uploaded_file(file_id)

                self.assertEqual(caught.exception.code, "UPLOAD_GONE")
                self.assertFalse(upload_dir.exists())

    def test_cleanup_removes_deterministic_upload_without_cache_entry(self):
        file_id, _ = self._store_csv()
        upload_dir = coding._upload_dir_for_id(file_id)
        coding._uploaded_files.clear()

        coding._cleanup_file_id(file_id)

        self.assertFalse(os.path.exists(upload_dir))

    def test_sweeper_removes_upload_without_cache_at_ttl_boundary(self):
        file_id, _ = self._store_csv()
        upload_dir = coding._upload_dir_for_id(file_id)
        directory_time = 1_000.0
        os.utime(upload_dir, (directory_time, directory_time))
        coding._uploaded_files.clear()

        with patch.object(
            coding.time,
            "time",
            return_value=directory_time + coding._TEMP_TTL_SECONDS,
        ):
            coding.sweep_temp_files()

        self.assertFalse(os.path.exists(upload_dir))

    def test_upload_route_creates_recoverable_upload_and_status_is_not_cached(self):
        response = self.client.post(
            "/api/coding/upload",
            files={"file": ("sample.csv", b"message,round\nhello,1\n", "text/csv")},
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["file_name"], "sample.csv")
        self.assertEqual(body["row_count"], 1)

        coding._uploaded_files.clear()
        status = self.client.get(f"/api/coding/upload-status/{body['file_id']}")

        self.assertEqual(status.status_code, 200)
        self.assertTrue(status.json()["ok"])
        self.assertEqual(status.json()["file_name"], "sample.csv")
        self.assertEqual(status.headers["cache-control"], "no-store")

    def test_upload_status_has_structured_404_and_410_responses(self):
        malformed = self.client.get("/api/coding/upload-status/not-an-upload-id")
        self.assertEqual(malformed.status_code, 404)
        self.assertEqual(malformed.json()["code"], "UPLOAD_NOT_FOUND")
        self.assertFalse(malformed.json()["ok"])
        self.assertEqual(malformed.headers["cache-control"], "no-store")

        gone_id = uuid.uuid4().hex
        gone = self.client.get(f"/api/coding/upload-status/{gone_id}")
        self.assertEqual(gone.status_code, 410)
        self.assertEqual(gone.json()["code"], "UPLOAD_GONE")
        self.assertFalse(gone.json()["ok"])
        self.assertEqual(gone.headers["x-chat-error-code"], "UPLOAD_GONE")

    def test_category_generator_recovers_optional_sample_after_cache_clear(self):
        file_id, _ = self._store_csv()
        coding._uploaded_files.clear()
        prompts: list[str] = []

        class FakeProvider:
            async def complete(self, prompt, **kwargs):
                prompts.append(prompt)
                return {"response": '[{"label":"cooperation"}]'}

        async def collect_updates():
            updates = []
            with (
                patch.object(category_runner, "_get_provider_instance", return_value=FakeProvider()),
                patch.object(category_runner.asyncio, "sleep", return_value=None),
            ):
                async for update in category_runner.run_category_generation(
                    provider="openai",
                    model="example-model",
                    api_key="placeholder",
                    goals="Find cooperation",
                    hypothesis="",
                    output_type="classify",
                    target_count=1,
                    domain="",
                    references="",
                    file_id=file_id,
                    message_column="message",
                ):
                    updates.append(update)
            return updates

        updates = asyncio.run(collect_updates())

        self.assertIn('"text": "hello"', prompts[0])
        self.assertEqual(updates[-1], {"type": "complete", "total": 1})

    def test_agreement_cross_check_recovers_raters_after_cache_clear(self):
        first_id, _ = coding._store_uploaded_file(
            b"episode,label\n1,yes\n2,no\n", "human.csv", "csv"
        )
        second_id, _ = coding._store_uploaded_file(
            b"episode,label\n1,yes\n2,no\n", "llm.csv", "csv"
        )
        coding._uploaded_files.clear()

        response = self.client.post(
            "/api/agreement/cross-check",
            json={
                "raters": [
                    {"file_id": first_id, "name": "Human", "rater_type": "human"},
                    {"file_id": second_id, "name": "LLM", "rater_type": "llm"},
                ],
                "episode_columns": ["episode"],
                "analysis_variables": ["label"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.assertEqual(response.json()["common_episodes"], 2)

    def test_run_stream_rejects_stale_upload_before_streaming_200(self):
        gone_id = uuid.uuid4().hex

        response = self.client.post(
            "/api/coding/run-stream",
            json={"file_id": gone_id},
        )

        self.assertEqual(response.status_code, 410)
        self.assertEqual(response.json()["code"], "UPLOAD_GONE")
        self.assertEqual(
            response.json()["detail"],
            "Uploaded dataset is no longer available. Please re-upload it.",
        )
        self.assertEqual(response.headers["x-chat-error-code"], "UPLOAD_GONE")
        self.assertNotEqual(response.headers.get("content-type"), "application/x-ndjson")

    def test_package_generation_recovers_upload_without_process_cache(self):
        file_id, _ = self._store_csv()
        coding._uploaded_files.clear()
        payload = {
            "file_id": file_id,
            "file_name": "sample.csv",
            "message_column": "message",
            "experiment_instructions": "Participants send one message.",
            "codebook": [{"label": "cooperation", "type": "binary"}],
            "provider": "openai",
            "model": "example-model",
            "api_key": "not-written-to-package",
        }

        response = self.client.post("/api/coding/generate-package", json=payload)

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content)) as package:
            names = set(package.namelist())
            self.assertIn("sample_chat_input.xlsx", names)
            self.assertIn("code_sample.py", names)
            self.assertIn("README.md", names)
            self.assertIn("requirements.txt", names)
            script = package.read("code_sample.py").decode("utf-8")
            self.assertIn("sample_chat_input.xlsx", script)
            self.assertIn("PACKAGE_SOURCE_SHEET = 'source_rows'", script)
            self.assertIn("PACKAGE_EPISODE_SHEET = 'episodes'", script)
            self.assertIn("PACKAGE_ROW_MAP_SHEET = 'row_map'", script)
            self.assertIn("--also-save-episodes", script)
            compile(script, "code_sample.py", "exec")
            self.assertNotIn("not-written-to-package", script)

            workbook = load_workbook(
                io.BytesIO(package.read("sample_chat_input.xlsx")),
                data_only=True,
            )
            self.assertEqual(
                workbook.sheetnames,
                ["source_rows", "episodes", "row_map"],
            )
            self.assertEqual(workbook["source_rows"].max_row - 1, 1)
            self.assertEqual(workbook["episodes"].max_row - 1, 1)
            self.assertEqual(
                list(workbook["row_map"].values),
                [("source_row", "episode"), (1, 1)],
            )
            requirements = package.read("requirements.txt").decode("utf-8")
            self.assertIn("openpyxl", requirements.splitlines())

    def test_package_generation_returns_stable_gone_upload_signal(self):
        payload = {
            "file_id": uuid.uuid4().hex,
            "file_name": "sample.csv",
            "message_column": "message",
            "experiment_instructions": "Participants send one message.",
            "codebook": [{"label": "cooperation", "type": "binary"}],
            "provider": "openai",
            "model": "example-model",
            "api_key": "placeholder",
        }

        response = self.client.post("/api/coding/generate-package", json=payload)

        self.assertEqual(response.status_code, 410)
        self.assertEqual(response.json()["code"], "UPLOAD_GONE")
        self.assertEqual(response.headers["x-chat-error-code"], "UPLOAD_GONE")
        self.assertIn("re-upload", response.json()["detail"].lower())

    def test_generated_package_expands_codes_and_makes_episode_file_optional(self):
        file_id, _ = coding._store_uploaded_file(
            (
                b"episode,turn,sender,message,condition,cooperation\n"
                b"A,2,P2,second,treatment,human\n"
                b"A,1,P1,first,treatment,human\n"
                b"B,1,P1,,control,human\n"
            ),
            "study.csv",
            "csv",
        )
        payload = {
            "file_id": file_id,
            "file_name": "study.csv",
            "message_column": "message",
            "identifier_columns": ["episode"],
            "identity_column": "sender",
            "order_column": "turn",
            "context": [{"column": "condition", "description": "study arm"}],
            "experiment_instructions": "Participants send messages.",
            "codebook": [{"label": "cooperation", "type": "binary"}],
            "provider": "openai",
            "model": "example-model",
            "api_key": "not-written-to-package",
            "empty_message_handling": "ignore",
        }

        response = self.client.post("/api/coding/generate-package", json=payload)

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content)) as package:
            script = package.read("code_study.py").decode("utf-8")
            input_path = Path(self.temp_dir.name) / "study_chat_input.xlsx"
            input_path.write_bytes(package.read("study_chat_input.xlsx"))

        namespace = {"__name__": "generated_chat_package"}
        with patch.dict(os.environ, {"CHAT_API_KEY": "test-key"}):
            exec(compile(script, "code_study.py", "exec"), namespace)
        source_df, episode_df, row_map = namespace["load_package_workbook"](
            str(input_path)
        )
        namespace["call_llm"] = lambda prompt: {"cooperation": 1}
        episode_codes = namespace["code_package_episodes"](episode_df)
        primary, compact = namespace["build_package_results"](
            source_df,
            episode_df,
            row_map,
            episode_codes,
        )

        self.assertEqual(len(primary), 3)
        self.assertEqual(primary["cooperation"].tolist(), ["human", "human", "human"])
        self.assertEqual(primary["cooperation_coded"].iloc[:2].tolist(), [1, 1])
        self.assertTrue(pd.isna(primary["cooperation_coded"].iloc[2]))
        self.assertEqual(len(compact), 2)
        self.assertEqual(compact.loc[0, "message"], "[P1] first\n[P2] second")
        self.assertTrue(pd.isna(compact.loc[1, "cooperation_coded"]))

        primary_path, compact_path = namespace["save_package_results"](
            primary,
            compact,
            str(input_path),
            False,
        )
        self.assertTrue(Path(primary_path).is_file())
        self.assertEqual(load_workbook(primary_path).sheetnames, ["Coded data"])
        self.assertIsNone(compact_path)
        self.assertFalse((Path(self.temp_dir.name) / "study_coded_episodes.xlsx").exists())

        _, compact_path = namespace["save_package_results"](
            primary,
            compact,
            str(input_path),
            True,
        )
        self.assertTrue(Path(compact_path).is_file())
        self.assertEqual(load_workbook(compact_path).sheetnames, ["Coded episodes"])

    def test_status_reports_expired_metadata_as_410_and_removes_it(self):
        file_id, _ = self._store_csv()
        metadata_path = (
            Path(coding._upload_dir_for_id(file_id))
            / coding._UPLOAD_METADATA_FILENAME
        )
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        metadata["created"] = time.time() - coding._TEMP_TTL_SECONDS
        metadata_path.write_text(json.dumps(metadata), encoding="utf-8")
        coding._uploaded_files.clear()

        response = self.client.get(f"/api/coding/upload-status/{file_id}")

        self.assertEqual(response.status_code, 410)
        self.assertEqual(response.json()["code"], "UPLOAD_EXPIRED")
        self.assertEqual(response.headers["cache-control"], "no-store")
        self.assertFalse(os.path.exists(coding._upload_dir_for_id(file_id)))


if __name__ == "__main__":
    unittest.main()
