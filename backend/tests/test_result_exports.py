import io
import os
import tempfile
import unittest
import zipfile
from unittest.mock import patch

import pandas as pd
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded

from app.ratelimit import limiter
from app.routes import coding
from app.services.result_exporter import (
    build_result_frames,
    dataframe_to_xlsx,
    detect_sender_names,
    expanded_codebook_labels,
    find_context_conflicts,
    prepare_coding_dataset,
    validate_sender_configuration,
)
from app.services.coding_runner import aggregate_output_labels, aggregate_results


class ResultFrameTests(unittest.TestCase):
    def setUp(self):
        self.source = pd.DataFrame(
            [
                {"session": "A", "turn": 2, "sender": "P2", "message": "second", "condition": "treatment"},
                {"session": "A", "turn": 1, "sender": "P1", "message": "first", "condition": "treatment"},
                {"session": "B", "turn": 1, "sender": "P1", "message": "only", "condition": "control"},
            ]
        )

    def _prepared(self):
        return prepare_coding_dataset(
            self.source,
            message_column="message",
            identifier_columns=["session"],
            identity_column="sender",
            order_column="turn",
            order_direction="asc",
        )

    def test_preprocessing_preserves_source_lineage_and_message_order(self):
        prepared = self._prepared()

        self.assertEqual(prepared.source_episode_indices, [0, 0, 1])
        self.assertEqual(len(prepared.episodes), 2)
        self.assertEqual(prepared.episodes.loc[0, "message"], "[P1] first\n[P2] second")
        self.assertEqual(prepared.episodes.loc[0, "sender"], "P1\nP2")
        self.assertEqual(prepared.episodes.loc[0, "turn"], "1\n2")
        self.assertEqual(prepared.episodes.loc[1, "message"], "[P1] only")

    def test_tied_order_values_preserve_uploaded_row_order(self):
        source = pd.DataFrame(
            [
                {"episode": "A", "turn": 1, "message": "first uploaded"},
                {"episode": "A", "turn": 1, "message": "second uploaded"},
                {"episode": "A", "turn": 2, "message": "later turn"},
            ]
        )

        ascending = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=["episode"],
            identity_column=None,
            order_column="turn",
            order_direction="asc",
        )
        descending = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=["episode"],
            identity_column=None,
            order_column="turn",
            order_direction="desc",
        )

        self.assertEqual(
            ascending.episodes.loc[0, "message"],
            "first uploaded\nsecond uploaded\nlater turn",
        )
        self.assertEqual(
            descending.episodes.loc[0, "message"],
            "later turn\nfirst uploaded\nsecond uploaded",
        )

    def test_context_must_match_exactly_within_each_episode(self):
        source = pd.DataFrame(
            [
                {"episode": "A", "message": "one", "condition": "control"},
                {"episode": "A", "message": "two", "condition": "treatment"},
                {"episode": "B", "message": "three", "condition": "control"},
                {"episode": "B", "message": "four", "condition": "control"},
            ]
        )

        conflicts = find_context_conflicts(
            source,
            identifier_columns=["episode"],
            context_columns=["condition"],
        )

        self.assertEqual(len(conflicts), 1)
        self.assertEqual(conflicts[0]["column"], "condition")
        self.assertEqual(conflicts[0]["conflicting_episode_count"], 1)
        self.assertEqual(conflicts[0]["example_episode"], {"episode": "A"})
        with self.assertRaisesRegex(ValueError, "exactly one value"):
            prepare_coding_dataset(
                source,
                message_column="message",
                identifier_columns=["episode"],
                identity_column=None,
                order_column=None,
                order_direction="asc",
                context_columns=["condition"],
            )

    def test_blank_and_nonblank_context_values_conflict(self):
        source = pd.DataFrame(
            [
                {"episode": "A", "message": "one", "condition": None},
                {"episode": "A", "message": "two", "condition": "control"},
            ]
        )

        conflicts = find_context_conflicts(
            source,
            identifier_columns=["episode"],
            context_columns=["condition"],
        )

        self.assertEqual(conflicts[0]["example_values"], ["(blank)", "control"])

    def test_sender_names_are_detected_and_must_match_verified_list(self):
        source = pd.DataFrame(
            [
                {"sender": "P2", "message": "one"},
                {"sender": "P1", "message": "two"},
                {"sender": "P2", "message": "three"},
            ]
        )
        codebook = [{"label": "tone", "level": "sender"}]

        names, blank_rows = detect_sender_names(source, "sender")

        self.assertEqual(names, ["P2", "P1"])
        self.assertEqual(blank_rows, [])
        self.assertEqual(
            validate_sender_configuration(
                source,
                identity_column="sender",
                participants=["P2", "P1"],
                codebook=codebook,
            ),
            ["P2", "P1"],
        )
        with self.assertRaisesRegex(ValueError, "Expected, in first-appearance order"):
            validate_sender_configuration(
                source,
                identity_column="sender",
                participants=["P1", "P2"],
                codebook=codebook,
            )

    def test_blank_sender_blocks_sender_level_coding(self):
        source = pd.DataFrame([{"sender": "P1"}, {"sender": None}])

        with self.assertRaisesRegex(ValueError, r"blank values in source row\(s\) 2"):
            validate_sender_configuration(
                source,
                identity_column="sender",
                participants=["P1"],
                codebook=[{"label": "tone", "level": "sender"}],
            )

    def test_noncontiguous_source_rows_use_positional_episode_mapping(self):
        source = pd.DataFrame(
            [
                {"episode": "A", "message": "first A"},
                {"episode": "B", "message": "only B"},
                {"episode": "A", "message": "second A"},
            ]
        )
        prepared = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=["episode"],
            identity_column=None,
            order_column=None,
            order_direction="asc",
        )
        primary, _, _ = build_result_frames(
            source_df=source,
            prepared=prepared,
            coded_rows=[
                {"index": 0, "coded": {"label": "A-code"}},
                {"index": 1, "coded": {"label": "B-code"}},
            ],
            codebook=[{"label": "label", "level": "episode"}],
            participants=[],
            message_column="message",
            identifier_columns=["episode"],
            context_columns=[],
        )

        self.assertEqual(prepared.source_episode_indices, [0, 1, 0])
        self.assertEqual(primary["label"].tolist(), ["A-code", "B-code", "A-code"])

    def test_row_as_episode_maps_each_source_row_to_itself(self):
        source = pd.DataFrame(
            [
                {"duplicate_id": 1, "message": "one"},
                {"duplicate_id": 1, "message": "two"},
            ]
        )
        prepared = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=[],
            identity_column=None,
            order_column=None,
            order_direction="asc",
        )

        self.assertEqual(prepared.source_episode_indices, [0, 1])
        pd.testing.assert_frame_equal(prepared.episodes, source)

    def test_primary_repeats_episode_codes_without_changing_source_rows(self):
        source_before = self.source.copy(deep=True)
        primary, compact, names = build_result_frames(
            source_df=self.source,
            prepared=self._prepared(),
            coded_rows=[
                {"index": 0, "coded": {"cooperation": 1, "tone_P1": "warm", "tone_P2": "neutral"}},
                {"index": 1, "coded": {"cooperation": 0, "tone_P1": "cold", "tone_P2": None}},
            ],
            codebook=[
                {"label": "cooperation", "level": "episode"},
                {"label": "tone", "level": "sender"},
            ],
            participants=["P1", "P2"],
            message_column="message",
            identifier_columns=["session"],
            context_columns=["condition"],
            identity_column="sender",
            order_column="turn",
        )

        pd.testing.assert_frame_equal(self.source, source_before)
        self.assertEqual(len(primary), len(self.source))
        self.assertEqual(
            list(primary.columns),
            [*self.source.columns, "cooperation", "tone_P1", "tone_P2"],
        )
        self.assertEqual(primary["cooperation"].tolist(), [1, 1, 0])
        self.assertEqual(primary["tone_P1"].tolist(), ["warm", "warm", "cold"])
        self.assertEqual(names["cooperation"], "cooperation")
        self.assertEqual(
            list(compact.columns),
            ["session", "message", "sender", "turn", "condition", "cooperation", "tone_P1", "tone_P2"],
        )
        self.assertEqual(len(compact), 2)

    def test_missing_ignored_episode_is_retained_with_blank_codes(self):
        source = pd.DataFrame(
            [
                {"episode": 1, "message": ""},
                {"episode": 1, "message": None},
                {"episode": 2, "message": "hello"},
            ]
        )
        prepared = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=["episode"],
            identity_column=None,
            order_column=None,
            order_direction="asc",
        )
        self.assertEqual(prepared.episodes.loc[0, "message"], "\n")

        primary, compact, _ = build_result_frames(
            source_df=source,
            prepared=prepared,
            coded_rows=[{"index": 1, "coded": {"label": "yes"}}],
            codebook=[{"label": "label", "level": "episode"}],
            participants=[],
            message_column="message",
            identifier_columns=["episode"],
            context_columns=[],
        )

        self.assertTrue(pd.isna(primary.loc[0, "label"]))
        self.assertTrue(pd.isna(primary.loc[1, "label"]))
        self.assertEqual(primary.loc[2, "label"], "yes")
        self.assertTrue(pd.isna(compact.loc[0, "label"]))

    def test_latest_duplicate_episode_result_wins(self):
        primary, _, _ = build_result_frames(
            source_df=self.source,
            prepared=self._prepared(),
            coded_rows=[
                {"index": 0, "coded": {"label": "old"}},
                {"index": 0, "coded": {"label": "new"}},
            ],
            codebook=[{"label": "label", "level": "episode"}],
            participants=[],
            message_column="message",
            identifier_columns=["session"],
            context_columns=[],
        )

        self.assertEqual(primary["label"].tolist()[:2], ["new", "new"])

    def test_original_column_collision_gets_unambiguous_coded_suffix(self):
        source = self.source.assign(cooperation=["human", "human", "human"])
        prepared = prepare_coding_dataset(
            source,
            message_column="message",
            identifier_columns=["session"],
            identity_column="sender",
            order_column="turn",
            order_direction="asc",
        )
        primary, _, names = build_result_frames(
            source_df=source,
            prepared=prepared,
            coded_rows=[{"index": 0, "coded": {"cooperation": 1}}],
            codebook=[{"label": "cooperation", "level": "episode"}],
            participants=[],
            message_column="message",
            identifier_columns=["session"],
            context_columns=[],
        )

        self.assertEqual(names["cooperation"], "cooperation_coded")
        self.assertEqual(primary["cooperation"].tolist(), ["human", "human", "human"])
        self.assertEqual(primary["cooperation_coded"].tolist()[:2], [1, 1])

    def test_duplicate_expanded_codebook_labels_are_rejected(self):
        with self.assertRaisesRegex(ValueError, "tone_P1"):
            expanded_codebook_labels(
                [
                    {"label": "tone_P1", "level": "episode"},
                    {"label": "tone", "level": "sender"},
                ],
                ["P1"],
            )

    def test_xlsx_writer_preserves_formula_like_text_as_data(self):
        content = dataframe_to_xlsx(
            pd.DataFrame([{"message": "=HYPERLINK(\"https://example.com\")", "code": 1}]),
            sheet_name="Coded data",
        )
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        cell = workbook["Coded data"]["A2"]

        self.assertEqual(cell.value, '=HYPERLINK("https://example.com")')
        self.assertEqual(cell.data_type, "s")

    def test_xlsx_writer_preserves_formula_like_header_as_data(self):
        content = dataframe_to_xlsx(
            pd.DataFrame([["value"]], columns=["=SUM(A1:A2)"]),
            sheet_name="Coded data",
        )
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        cell = workbook["Coded data"]["A1"]

        self.assertEqual(cell.value, "=SUM(A1:A2)")
        self.assertEqual(cell.data_type, "s")

    def test_xlsx_writer_rejects_oversized_string_instead_of_truncating(self):
        with self.assertRaises(ValueError) as raised:
            dataframe_to_xlsx(
                pd.DataFrame([{"message": "x" * 32_768}]),
                sheet_name="Coded data",
            )

        message = str(raised.exception)
        self.assertIn("cell A2", message)
        self.assertIn("32,768 characters", message)
        self.assertIn("at most 32,767", message)

    def test_xlsx_writer_rejects_xml_illegal_control_character(self):
        with self.assertRaises(ValueError) as raised:
            dataframe_to_xlsx(
                pd.DataFrame([{"message": "before\u000bafter"}]),
                sheet_name="Coded data",
            )

        message = str(raised.exception)
        self.assertIn("cell A2", message)
        self.assertIn("XML-incompatible", message)
        self.assertIn("U+000B", message)


class AggregationTests(unittest.TestCase):
    def test_numeric_mode_uses_median_when_there_is_no_unique_mode(self):
        codebook = [
            {"label": "score", "type": "numeric", "level": "episode", "aggregation": "mode"}
        ]

        result = aggregate_results(
            [{"score": 1}, {"score": 2}, {"score": 3}, {"score": 4}],
            codebook,
            [],
        )

        self.assertEqual(result, {"score": 2.5})

    def test_numeric_mode_retains_a_unique_mode(self):
        codebook = [
            {"label": "score", "type": "numeric", "level": "episode", "aggregation": "mode"}
        ]

        result = aggregate_results(
            [{"score": 2}, {"score": 2}, {"score": 9}, {"score": 10}],
            codebook,
            [],
        )

        self.assertEqual(result, {"score": 2.0})

    def test_categorical_values_expand_to_independently_aggregated_binary_columns(self):
        codebook = [{
            "label": "option",
            "type": "categorical",
            "level": "episode",
            "aggregation": "mean",
            "values": [{"value": "a"}, {"value": "b"}, {"value": "c"}],
        }]

        result = aggregate_results(
            [{"option": "a"}, {"option": "a"}, {"option": "b"}, {"option": "c"}],
            codebook,
            [],
        )

        self.assertEqual(result, {"option_a": 0.5, "option_b": 0.25, "option_c": 0.25})

    def test_categorical_mode_tie_uses_binary_column_medians(self):
        codebook = [{
            "label": "option",
            "type": "categorical",
            "level": "episode",
            "aggregation": "mode",
            "values": [{"value": "a"}, {"value": "b"}],
        }]

        result = aggregate_results(
            [{"option": "a"}, {"option": "b"}],
            codebook,
            [],
        )

        self.assertEqual(result, {"option_a": 0.5, "option_b": 0.5})

    def test_text_is_excluded_from_aggregate_output(self):
        codebook = [
            {"label": "explanation", "type": "text", "level": "episode", "aggregation": "mode"},
            {"label": "score", "type": "numeric", "level": "episode", "aggregation": "mean"},
        ]

        self.assertEqual(aggregate_output_labels(codebook, []), ["score"])
        self.assertEqual(
            aggregate_results([{"explanation": "one", "score": 2}], codebook, []),
            {"score": 2.0},
        )

    def test_sanitized_categorical_values_must_create_unique_columns(self):
        codebook = [{
            "label": "option",
            "type": "categorical",
            "level": "episode",
            "values": [{"value": "a b"}, {"value": "a/b"}],
        }]

        with self.assertRaisesRegex(ValueError, "option_a_b"):
            aggregate_output_labels(codebook, [])


class ResultExportEndpointTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.tempdir_patch = patch.object(
            coding.tempfile,
            "gettempdir",
            return_value=self.temp_dir.name,
        )
        self.tempdir_patch.start()
        coding._uploaded_files.clear()

        app = FastAPI()
        app.state.limiter = limiter
        app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
        app.include_router(coding.router, prefix="/api")
        self.client = TestClient(app)

    def tearDown(self):
        self.client.close()
        coding._uploaded_files.clear()
        self.tempdir_patch.stop()
        self.temp_dir.cleanup()

    def _store_dataset(self):
        content = (
            b"session,turn,sender,message,condition\n"
            b"A,2,P2,second,treatment\n"
            b"A,1,P1,first,treatment\n"
            b"B,1,P1,only,control\n"
        )
        return coding._store_uploaded_file(content, "study.csv", "csv")[0]

    def _payload(self, kind: str):
        return {
            "file_id": self._store_dataset(),
            "message_column": "message",
            "identifier_columns": ["session"],
            "identity_column": "sender",
            "order_column": "turn",
            "order_direction": "asc",
            "context": [{"column": "condition", "description": "treatment arm"}],
            "codebook": [{"label": "cooperation", "type": "binary"}],
            "participants": [],
            "coded_rows": [
                {"index": 0, "coded": {"cooperation": 1}},
                {"index": 1, "coded": {"cooperation": 0}},
            ],
            "kind": kind,
        }

    def test_primary_endpoint_returns_original_rows_in_csv(self):
        response = self.client.post("/api/coding/export-results", json=self._payload("primary"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("study_coded.csv", response.headers["content-disposition"])
        self.assertTrue(response.headers["content-type"].startswith("text/csv"))
        values = [tuple(row) for row in pd.read_csv(io.BytesIO(response.content)).itertuples(index=False, name=None)]
        self.assertEqual(
            tuple(pd.read_csv(io.BytesIO(response.content), nrows=0).columns),
            ("session", "turn", "sender", "message", "condition", "cooperation"),
        )
        self.assertEqual(len(values), 3)
        self.assertEqual([row[-1] for row in values], [1, 1, 0])

    def test_episode_endpoint_contains_only_compact_mapped_columns(self):
        response = self.client.post("/api/coding/export-results", json=self._payload("episodes"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("study_coded_episodes.csv", response.headers["content-disposition"])
        frame = pd.read_csv(io.BytesIO(response.content))
        values = [tuple(row) for row in frame.itertuples(index=False, name=None)]
        self.assertEqual(
            tuple(frame.columns),
            ("session", "message", "sender", "turn", "condition", "cooperation"),
        )
        self.assertEqual(values[0][1], "[P1] first\n[P2] second")
        self.assertEqual(values[0][2], "P1\nP2")
        self.assertEqual(values[0][3], "1\n2")
        self.assertEqual(len(values), 2)

    def test_export_endpoint_rejects_duplicate_expanded_labels(self):
        payload = self._payload("primary")
        payload["codebook"] = [
            {"label": "tone_P1", "type": "text", "level": "episode"},
            {"label": "tone", "type": "text", "level": "sender"},
        ]
        payload["participants"] = ["P2", "P1"]

        response = self.client.post("/api/coding/export-results", json=payload)

        self.assertEqual(response.status_code, 400)
        self.assertIn("tone_P1", response.json()["detail"])

    def test_csv_export_is_not_limited_by_excel_cell_length(self):
        payload = self._payload("primary")
        payload["coded_rows"][0]["coded"]["cooperation"] = "x" * 32_768

        response = self.client.post("/api/coding/export-results", json=payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"x" * 32_768, response.content)

    def test_aggregated_download_contains_overall_model_and_original_run_files(self):
        payload = self._payload("primary")
        payload["codebook"] = [
            {"label": "cooperation", "type": "binary", "level": "episode", "aggregation": "mode"},
            {
                "label": "option",
                "type": "categorical",
                "level": "episode",
                "aggregation": "mean",
                "values": [{"value": "a"}, {"value": "b"}],
            },
            {"label": "score", "type": "numeric", "level": "episode", "aggregation": "mode"},
            {"label": "note", "type": "text", "level": "episode", "aggregation": "mode"},
        ]
        payload["model_call_count"] = 4
        payload["coded_rows"] = [
            {
                "index": 0,
                "coded": {
                    "cooperation": 1,
                    "option_a": 0.75,
                    "option_b": 0.25,
                    "score": 4,
                },
            },
            {
                "index": 1,
                "coded": {
                    "cooperation": 0.5,
                    "option_a": 0.5,
                    "option_b": 0.5,
                    "score": 2.5,
                },
            },
        ]

        detail_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        detail_path = os.path.join(detail_dir, "coded_results.csv")
        records = []
        episode_data = {
            0: {"session": "A", "turn": "1\n2", "sender": "P1\nP2", "message": "[P1] first\n[P2] second", "condition": "treatment"},
            1: {"session": "B", "turn": "1", "sender": "P1", "message": "[P1] only", "condition": "control"},
        }
        calls = {
            0: [
                ("openai/model__run1", 1, "a", 1, "oa1"),
                ("openai/model__run2", 1, "b", 3, "oa2"),
                ("gemini/model__run1", 0, "a", 5, "ga1"),
                ("gemini/model__run2", 1, "a", 7, "ga2"),
            ],
            1: [
                ("openai/model__run1", 0, "a", 1, "ob1"),
                ("openai/model__run2", 1, "b", 4, "ob2"),
                ("gemini/model__run1", 0, "a", 2, "gb1"),
                ("gemini/model__run2", 1, "b", 3, "gb2"),
            ],
        }
        for episode_index, episode_calls in calls.items():
            for coder, cooperation, option, score, note in episode_calls:
                records.append({
                    **episode_data[episode_index],
                    "__chat_episode_index": episode_index,
                    "coder": coder,
                    "cooperation": cooperation,
                    "option": option,
                    "score": score,
                    "note": note,
                })
        pd.DataFrame(records).to_csv(detail_path, index=False)
        payload["result_path"] = detail_path

        response = self.client.post("/api/coding/export-results", json=payload)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.headers["content-type"].startswith("application/zip"))
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                sorted([
                    "inter_coder_agreement.csv",
                    "overall/aggregated_results.csv",
                    "overall/text_results.csv",
                    "models/openai_model/aggregated_results.csv",
                    "models/openai_model/text_results.csv",
                    "models/openai_model/runs/run1.csv",
                    "models/openai_model/runs/run2.csv",
                    "models/gemini_model/aggregated_results.csv",
                    "models/gemini_model/text_results.csv",
                    "models/gemini_model/runs/run1.csv",
                    "models/gemini_model/runs/run2.csv",
                ]),
            )

            overall = pd.read_csv(io.BytesIO(archive.read("overall/aggregated_results.csv")))
            overall_text = pd.read_csv(io.BytesIO(archive.read("overall/text_results.csv")))
            openai_aggregate = pd.read_csv(
                io.BytesIO(archive.read("models/openai_model/aggregated_results.csv"))
            )
            openai_run_one = pd.read_csv(
                io.BytesIO(archive.read("models/openai_model/runs/run1.csv"))
            )
            agreement = pd.read_csv(
                io.BytesIO(archive.read("inter_coder_agreement.csv"))
            )

        self.assertEqual(len(overall), 3)
        self.assertEqual(
            list(overall.columns[-4:]),
            ["cooperation", "option_a", "option_b", "score"],
        )
        self.assertEqual(overall["score"].tolist(), [4.0, 4.0, 2.5])
        self.assertNotIn("note", overall.columns)
        self.assertEqual(len(overall_text), 8)
        self.assertEqual(set(overall_text["note"]), {"oa1", "oa2", "ga1", "ga2", "ob1", "ob2", "gb1", "gb2"})
        self.assertEqual(openai_aggregate["score"].tolist(), [2.0, 2.5])
        self.assertEqual(openai_aggregate["option_a"].tolist(), [0.5, 0.5])
        self.assertEqual(openai_run_one["option"].tolist(), ["a", "a"])
        self.assertEqual(openai_run_one["note"].tolist(), ["oa1", "ob1"])
        self.assertEqual(
            list(agreement.columns),
            ["model_a", "model_b", "variable", "agreement_rate", "cohens_kappa", "paired_n"],
        )
        self.assertEqual(set(agreement["model_a"]), {"openai/model"})
        self.assertEqual(set(agreement["model_b"]), {"gemini/model"})
        self.assertIn("cooperation", set(agreement["variable"]))
        self.assertNotIn("AVERAGE", set(agreement["variable"]))

    def test_inter_coder_endpoint_aggregates_runs_within_models(self):
        detail_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        detail_path = os.path.join(detail_dir, "coded_results.csv")
        pd.DataFrame([
            {"__chat_episode_index": 0, "coder": "openai/model__run1", "cooperation": 1},
            {"__chat_episode_index": 0, "coder": "openai/model__run2", "cooperation": 1},
            {"__chat_episode_index": 0, "coder": "gemini/model__run1", "cooperation": 1},
            {"__chat_episode_index": 0, "coder": "gemini/model__run2", "cooperation": 0},
            {"__chat_episode_index": 1, "coder": "openai/model__run1", "cooperation": 0},
            {"__chat_episode_index": 1, "coder": "openai/model__run2", "cooperation": 0},
            {"__chat_episode_index": 1, "coder": "gemini/model__run1", "cooperation": 1},
            {"__chat_episode_index": 1, "coder": "gemini/model__run2", "cooperation": 1},
        ]).to_csv(detail_path, index=False)

        response = self.client.post(
            "/api/coding/inter-coder-agreement",
            json={
                "result_path": detail_path,
                "codebook": [{
                    "label": "cooperation",
                    "type": "binary",
                    "level": "episode",
                    "aggregation": "mode",
                }],
                "participants": [],
            },
        )

        self.assertEqual(response.status_code, 200)
        report = response.json()
        self.assertTrue(report["eligible"])
        self.assertEqual(report["models"], ["openai/model", "gemini/model"])
        self.assertEqual(report["pairs"][0]["variables"][0]["n"], 2)

    def test_inter_coder_endpoint_rejects_unrelated_paths(self):
        unrelated = os.path.join(self.temp_dir.name, "unrelated.csv")
        pd.DataFrame([{"coder": "model", "value": 1}]).to_csv(unrelated, index=False)

        response = self.client.post(
            "/api/coding/inter-coder-agreement",
            json={
                "result_path": unrelated,
                "codebook": [{"label": "value", "type": "numeric"}],
                "participants": [],
            },
        )

        self.assertEqual(response.status_code, 403)

    def test_text_only_aggregated_download_omits_aggregate_csv_files(self):
        payload = self._payload("primary")
        payload["codebook"] = [
            {"label": "note", "type": "text", "level": "episode", "aggregation": "mode"}
        ]
        payload["model_call_count"] = 2
        payload["coded_rows"] = [{"index": 0, "coded": {}}, {"index": 1, "coded": {}}]
        detail_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        detail_path = os.path.join(detail_dir, "coded_results.csv")
        pd.DataFrame([
            {"__chat_episode_index": 0, "message": "first", "coder": "openai/model__run1", "note": "one"},
            {"__chat_episode_index": 0, "message": "first", "coder": "openai/model__run2", "note": "two"},
        ]).to_csv(detail_path, index=False)
        payload["result_path"] = detail_path

        response = self.client.post("/api/coding/export-results", json=payload)

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                sorted([
                    "overall/text_results.csv",
                    "models/openai_model/text_results.csv",
                    "models/openai_model/runs/run1.csv",
                    "models/openai_model/runs/run2.csv",
                ]),
            )

    def test_nontext_only_aggregated_download_omits_text_csv_files(self):
        payload = self._payload("primary")
        payload["model_call_count"] = 2
        detail_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        detail_path = os.path.join(detail_dir, "coded_results.csv")
        pd.DataFrame([
            {"__chat_episode_index": 0, "message": "first", "coder": "openai/model__run1", "cooperation": 1},
            {"__chat_episode_index": 0, "message": "first", "coder": "openai/model__run2", "cooperation": 0},
        ]).to_csv(detail_path, index=False)
        payload["result_path"] = detail_path

        response = self.client.post("/api/coding/export-results", json=payload)

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                sorted([
                    "overall/aggregated_results.csv",
                    "models/openai_model/aggregated_results.csv",
                    "models/openai_model/runs/run1.csv",
                    "models/openai_model/runs/run2.csv",
                ]),
            )

    def test_detailed_download_is_restricted_to_coding_result_artifacts(self):
        unrelated = os.path.join(self.temp_dir.name, "unrelated.csv")
        with open(unrelated, "w", encoding="utf-8") as handle:
            handle.write("private\nvalue\n")

        rejected = self.client.get("/api/coding/download", params={"path": unrelated})

        self.assertEqual(rejected.status_code, 403)

        result_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        result_path = os.path.join(result_dir, "coded_results.csv")
        with open(result_path, "w", encoding="utf-8") as handle:
            handle.write("message,label\nhello,yes\n")

        accepted = self.client.get("/api/coding/download", params={"path": result_path})

        self.assertEqual(accepted.status_code, 200)
        self.assertIn("coded_results.csv", accepted.headers["content-disposition"])

    def test_repeated_all_null_detailed_runs_download_without_mode_crash(self):
        result_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        result_path = os.path.join(result_dir, "coded_results.csv")
        with open(result_path, "w", encoding="utf-8") as handle:
            handle.write(
                "message,coder,score,_error\n"
                "hello,openai/model__run1,,api_failed\n"
                "hello,openai/model__run2,,api_failed\n"
            )

        response = self.client.get("/api/coding/download", params={"path": result_path})

        self.assertEqual(response.status_code, 200)
        self.assertIn("coded_results.zip", response.headers["content-disposition"])
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                ["openai_model/run1.csv", "openai_model/run2.csv"],
            )

    def test_selective_rerun_replaces_affected_detailed_records(self):
        index_column = "__chat_episode_index"
        previous_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        previous_path = os.path.join(previous_dir, "coded_results.csv")
        replacement_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        replacement_path = os.path.join(replacement_dir, "coded_results.csv")

        pd.DataFrame(
            [
                {index_column: 0, "message": "first", "coder": "openai/model__run1", "score": "old-0-1"},
                {index_column: 0, "message": "first", "coder": "openai/model__run2", "score": "old-0-2"},
                {index_column: 0, "message": "first", "coder": "__aggregated (per-variable)", "score": "old-0-a"},
                {index_column: 1, "message": "second", "coder": "openai/model__run1", "score": "old-1-1"},
                {index_column: 1, "message": "second", "coder": "openai/model__run2", "score": "old-1-2"},
                {index_column: 1, "message": "second", "coder": "__aggregated (per-variable)", "score": "old-1-a"},
            ]
        ).to_csv(previous_path, index=False)
        pd.DataFrame(
            [
                {index_column: 1, "message": "second", "coder": "openai/model__run1", "score": "new-1-1"},
                {index_column: 1, "message": "second", "coder": "openai/model__run2", "score": "new-1-2"},
                {index_column: 1, "message": "second", "coder": "__aggregated (per-variable)", "score": "new-1-a"},
            ]
        ).to_csv(replacement_path, index=False)

        merged_path = coding._merge_selective_rerun_artifact(
            previous_path,
            replacement_path,
            [1],
        )

        merged = pd.read_csv(merged_path)
        self.assertFalse(os.path.exists(previous_dir))
        self.assertEqual(len(merged), 6)
        self.assertEqual(
            merged.loc[merged[index_column] == 0, "score"].tolist(),
            ["old-0-1", "old-0-2", "old-0-a"],
        )
        self.assertEqual(
            merged.loc[merged[index_column] == 1, "score"].tolist(),
            ["new-1-1", "new-1-2", "new-1-a"],
        )
        self.assertFalse(merged["score"].str.startswith("old-1-").any())

        response = self.client.get("/api/coding/download", params={"path": merged_path})

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            aggregate = pd.read_csv(io.BytesIO(archive.read("aggregate.csv")))
            run_one = pd.read_csv(io.BytesIO(archive.read("openai_model/run1.csv")))
        self.assertNotIn(index_column, aggregate.columns)
        self.assertNotIn(index_column, run_one.columns)
        self.assertEqual(aggregate["score"].tolist(), ["old-0-a", "new-1-a"])
        self.assertEqual(run_one["score"].tolist(), ["old-0-1", "new-1-1"])

    def test_selective_rerun_requires_episode_indexed_previous_artifact(self):
        previous_dir = tempfile.mkdtemp(prefix="llm_coding_", dir=self.temp_dir.name)
        previous_path = os.path.join(previous_dir, "coded_results.csv")
        pd.DataFrame([{"message": "old", "coder": "model", "score": 1}]).to_csv(
            previous_path,
            index=False,
        )

        with self.assertRaisesRegex(ValueError, "Re-run all episodes once"):
            coding._validate_selective_rerun_artifact(previous_path)


if __name__ == "__main__":
    unittest.main()
